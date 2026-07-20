from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import urllib.error
import urllib.request
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


BEIJING_TZ = ZoneInfo("Asia/Shanghai")
REQUIRED_TEMPLATE_TOKENS = (
    "{{PERIOD_TYPE}}",
    "{{PERIOD_SLUG}}",
    "{{WECHAT_SUMMARY}}",
    "{{NEWS_SUMMARY}}",
)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate cross-repository AI trend reports")
    parser.add_argument("--period", choices=["week", "month", "both"], required=True)
    parser.add_argument("--week", default="last-week", help="last-week or YYYY-Www")
    parser.add_argument("--month", default="last-month", help="last-month or YYYY-MM")
    parser.add_argument("--wechat-root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--news-root", type=Path, default=Path("_sources/news-spider"))
    parser.add_argument(
        "--template",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "templates/cross-repo-trend-prompt.md",
    )
    parser.add_argument("--dry-run", action="store_true", help="Validate inputs without calling Z.AI")
    return parser.parse_args(argv)


def resolve_week(value: str, today: date | None = None) -> tuple[date, date, str]:
    today = today or datetime.now(BEIJING_TZ).date()
    if value in {"last-week", "previous-week"}:
        start = today - timedelta(days=today.weekday() + 7)
        iso_year, iso_week, _ = start.isocalendar()
        slug = f"{iso_year}-W{iso_week:02d}"
    else:
        try:
            year_text, week_text = value.split("-W", 1)
            start = date.fromisocalendar(int(year_text), int(week_text), 1)
        except (ValueError, TypeError) as exc:
            raise ValueError("week must be last-week or YYYY-Www") from exc
        slug = f"{start.isocalendar().year}-W{start.isocalendar().week:02d}"
    return start, start + timedelta(days=6), slug


def resolve_month(value: str, today: date | None = None) -> tuple[date, date, str]:
    today = today or datetime.now(BEIJING_TZ).date()
    if value in {"last-month", "previous-month"}:
        end = today.replace(day=1) - timedelta(days=1)
        start = end.replace(day=1)
    else:
        try:
            start = datetime.strptime(value, "%Y-%m").date().replace(day=1)
        except ValueError as exc:
            raise ValueError("month must be last-month or YYYY-MM") from exc
        if start.month == 12:
            next_month = date(start.year + 1, 1, 1)
        else:
            next_month = date(start.year, start.month + 1, 1)
        end = next_month - timedelta(days=1)
    return start, end, start.strftime("%Y-%m")


def compact_wechat_report(path: Path, max_chars: int = 60_000) -> str:
    if not path.exists():
        raise FileNotFoundError(f"WeChat summary not found: {path}")
    text = path.read_text(encoding="utf-8")
    archive_marker = "\n## 按日期归档"
    if archive_marker in text:
        text = text.split(archive_marker, 1)[0]
    text = text.strip()
    if len(text) > max_chars:
        text = text[:max_chars].rstrip() + "\n\n[微信公众号汇总内容已按字符上限截断]"
    return text


def make_excerpt(value: object, max_chars: int) -> str:
    normalized = " ".join(str(value or "").split())
    if len(normalized) <= max_chars:
        return normalized
    return normalized[:max_chars].rstrip() + "..."


def evenly_sample(rows: list[dict[str, str]], limit: int) -> list[dict[str, str]]:
    if len(rows) <= limit:
        return rows
    if limit <= 1:
        return rows[:limit]
    indices = {round(index * (len(rows) - 1) / (limit - 1)) for index in range(limit)}
    return [rows[index] for index in sorted(indices)]


def format_news_rows(rows: list[dict[str, str]], max_rows: int = 180) -> str:
    source_counter = Counter(row.get("source_name") or "未知来源" for row in rows)
    domain_counter = Counter(row.get("domain") or "未分类" for row in rows)
    selected = evenly_sample(rows, max_rows)
    lines = [
        f"文章总数：{len(rows)}",
        "主要来源：" + "、".join(f"{name}({count})" for name, count in source_counter.most_common(20)),
        "主要领域：" + "、".join(f"{name}({count})" for name, count in domain_counter.most_common(20)),
        f"代表性文章：{len(selected)} 篇",
        "",
    ]
    for index, row in enumerate(selected, 1):
        lines.append(
            f"{index}. [{row.get('published_at', '')}] "
            f"[{row.get('source_name', '')}｜{row.get('domain', '')}｜{row.get('sub_domain', '')}] "
            f"{row.get('title', '')}｜{make_excerpt(row.get('content', ''), 260)}"
        )
    return "\n".join(lines).strip()


def load_news_weekly(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"news-spider weekly summary not found: {path}")
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = [dict(row) for row in csv.DictReader(handle)]
    if not rows:
        raise ValueError(f"news-spider weekly summary is empty: {path}")
    return format_news_rows(rows)


def load_news_monthly(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"news-spider monthly summary not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Articles" not in workbook.sheetnames:
            raise ValueError(f"Monthly workbook has no Articles sheet: {path}")
        sheet = workbook["Articles"]
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values)]
        rows = [
            {headers[index]: str(value or "") for index, value in enumerate(row)}
            for row in values
        ]
    finally:
        workbook.close()
    if not rows:
        raise ValueError(f"news-spider monthly summary is empty: {path}")
    return format_news_rows(rows)


def render_prompt(template_path: Path, replacements: dict[str, str]) -> str:
    if not template_path.exists():
        raise FileNotFoundError(f"Trend prompt template not found: {template_path}")
    template = template_path.read_text(encoding="utf-8")
    missing_tokens = [token for token in REQUIRED_TEMPLATE_TOKENS if token not in template]
    if missing_tokens:
        raise ValueError("Template is missing required tokens: " + ", ".join(missing_tokens))
    for token, value in replacements.items():
        template = template.replace(token, value)
    return template.strip()


def extract_zai_text(payload: dict) -> str:
    choices = payload.get("choices") or []
    if not choices or not isinstance(choices[0], dict):
        return ""
    message = choices[0].get("message") or {}
    content = message.get("content") if isinstance(message, dict) else None
    return content.strip() if isinstance(content, str) else ""


def call_zai(prompt: str) -> tuple[str, str]:
    api_key = os.getenv("ZAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("ZAI_API_KEY is required for cross-repository AI trend reports")
    base_url = os.getenv("ZAI_BASE_URL", "").strip().rstrip("/") or "https://api.z.ai/api/paas/v4"
    model = os.getenv("ZAI_MODEL", "").strip() or "glm-5.2"
    payload = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": "你是一名严谨的中文产业情报分析师，只能依据用户提供的两个项目汇总材料进行分析。",
            },
            {"role": "user", "content": prompt},
        ],
        "thinking": {"type": "disabled"},
        "temperature": 0.3,
        "max_tokens": 5000,
        "stream": False,
    }
    request = urllib.request.Request(
        f"{base_url}/chat/completions",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=300) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"Z.AI trend request failed: {exc}") from exc
    text = extract_zai_text(response_payload)
    if not text:
        raise RuntimeError("Z.AI response did not contain choices[0].message.content")
    return text, model


def source_paths(
    period: str,
    slug: str,
    wechat_root: Path,
    news_root: Path,
) -> tuple[Path, Path, Path]:
    if period == "week":
        return (
            wechat_root / "reports" / "weekly" / f"{slug}.md",
            news_root / "data" / "weekly" / f"articles_week_{slug}.csv",
            wechat_root / "reports" / "trends" / "weekly" / f"{slug}.md",
        )
    return (
        wechat_root / "reports" / "monthly" / f"{slug}.md",
        news_root / "data" / "monthly" / f"articles_month_{slug}.xlsx",
        wechat_root / "reports" / "trends" / "monthly" / f"{slug}.md",
    )


def generate_one(
    period: str,
    slug: str,
    start: date,
    end: date,
    wechat_root: Path,
    news_root: Path,
    template_path: Path,
    dry_run: bool = False,
) -> Path | None:
    wechat_path, news_path, output_path = source_paths(period, slug, wechat_root, news_root)
    wechat_summary = compact_wechat_report(wechat_path)
    news_summary = load_news_weekly(news_path) if period == "week" else load_news_monthly(news_path)
    prompt = render_prompt(
        template_path,
        {
            "{{PERIOD_TYPE}}": "周报" if period == "week" else "月报",
            "{{PERIOD_SLUG}}": slug,
            "{{WECHAT_SUMMARY}}": wechat_summary,
            "{{NEWS_SUMMARY}}": news_summary,
        },
    )
    if dry_run:
        print(
            f"Validated {period} {slug}: wechat={wechat_path}, news={news_path}, "
            f"prompt_chars={len(prompt)}"
        )
        return None

    analysis, model = call_zai(prompt)
    period_label = "周度" if period == "week" else "月度"
    lines = [
        f"# 跨来源 AI 趋势{period_label}报告：{slug}",
        "",
        f"- 时间范围：{start.isoformat()} 至 {end.isoformat()}",
        f"- 输入项目：wechat-daily-news-csv、news-spider",
        f"- 模型：Z.AI `{model}`",
        "",
        "## 整体趋势研判",
        "",
        analysis.strip(),
        "",
        "## 输入文件",
        "",
        f"- `wechat-daily-news-csv/{wechat_path.relative_to(wechat_root).as_posix()}`",
        f"- `news-spider/{news_path.relative_to(news_root).as_posix()}`",
        "",
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")
    return output_path


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    wechat_root = args.wechat_root.resolve()
    news_root = args.news_root.resolve()
    template_path = args.template.resolve()
    requested = ["week", "month"] if args.period == "both" else [args.period]

    for period in requested:
        if period == "week":
            start, end, slug = resolve_week(args.week)
        else:
            start, end, slug = resolve_month(args.month)
        output = generate_one(
            period,
            slug,
            start,
            end,
            wechat_root,
            news_root,
            template_path,
            dry_run=args.dry_run,
        )
        if output:
            print(f"Generated cross-repository trend report: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
